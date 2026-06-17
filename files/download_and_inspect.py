import urllib.request
import openpyxl
import os

SPREADSHEETS = {
    "TTCS_2026": "1r4OsKbVzleV_eYGYEoAZxD1SEIezM6X289xP4ru1TK0",
    "TTCS_2025": "1DnsGIOTl23R3oRBbi1SuE-Sxh5cChxGKCR-kWLrBFVo",
    "TTTK_2025": "1JTwnQx1NzB2QJJ-njI3GY8gksPoYX4kFOGsadsEvVww",
    "TTTK_2026": "1byrhdkbpzxhg8dRKM1of7FurdogMdvQ-F5FpISUSM9E",
    "TTCXK_2025": "1iGs3bBIDdvf3TUiOxrkWNddwZuR63391hfnF8-OC20s",
    "TTCXK_2026": "1670sIuSotKrIH0sEb_sabGR3SlhwCm8QdK2f_mNPUwU"
}

def download_and_inspect():
    for name, sheet_id in SPREADSHEETS.items():
        url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
        filename = f"{name}.xlsx"
        print(f"\nDownloading {name} from {url}...")
        try:
            urllib.request.urlretrieve(url, filename)
            print(f"Downloaded {filename}")
            
            # Load with openpyxl
            wb = openpyxl.load_workbook(filename, read_only=True)
            print(f"Tabs in {name}: {wb.sheetnames}")
            
            # Print first 2 rows of each sheet
            for sheetname in wb.sheetnames:
                # Limit to first few sheets to avoid too much output
                sheet = wb[sheetname]
                rows = list(sheet.iter_rows(max_row=3, values_only=True))
                print(f"  Tab '{sheetname}':")
                for i, row in enumerate(rows):
                    # print non-empty fields or truncated row
                    clean_row = [str(cell)[:30] if cell is not None else "" for cell in row[:15]]
                    print(f"    Row {i+1}: {clean_row}")
            
            # Clean up
            os.remove(filename)
        except Exception as e:
            print(f"Error processing {name}: {e}")

if __name__ == "__main__":
    download_and_inspect()
