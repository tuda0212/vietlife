import urllib.request
import openpyxl
import os

def main():
    sheet_id = "1guUru-qTB6Pug4Oc43fqoEw3UUVmFXs2R7MAUr3AIGY"
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    filename = "user_sheet.xlsx"
    print(f"Downloading {url}...")
    try:
        urllib.request.urlretrieve(url, filename)
        print(f"Downloaded {filename}")
        
        wb = openpyxl.load_workbook(filename, data_only=True)
        print(f"Tabs in workbook: {wb.sheetnames}")
        
        for name in wb.sheetnames:
            sheet = wb[name]
            print(f"\nTab: {name} (Rows: {sheet.max_row}, Cols: {sheet.max_column})")
            # In ra 10 dòng đầu
            rows = list(sheet.iter_rows(max_row=20, values_only=True))
            for i, r in enumerate(rows):
                # Loại bỏ cell None ở cuối dòng để hiển thị gọn hơn
                non_empty = [str(x) if x is not None else "" for x in r]
                # Chỉ hiển thị nếu có dữ liệu
                if any(non_empty):
                    print(f"  Row {i+1}: {non_empty[:15]}")
                    
        os.remove(filename)
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    main()
