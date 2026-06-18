import urllib.request
import urllib.error
import subprocess
import openpyxl
import os

def get_access_token():
    try:
        token = subprocess.check_output(
            ["gcloud", "auth", "print-access-token"],
            text=True,
            stderr=subprocess.DEVNULL
        ).strip()
        return token
    except Exception as e:
        print(f"Error getting gcloud token: {e}")
        return None

def main():
    sheet_id = "1guUru-qTB6Pug4Oc43fqoEw3UUVmFXs2R7MAUr3AIGY"
    token = get_access_token()
    if not token:
        print("Could not retrieve access token.")
        return
        
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    filename = "user_sheet_auth.xlsx"
    
    print(f"Downloading spreadsheet via export link with auth token...")
    req = urllib.request.Request(url)
    req.add_header("Authorization", f"Bearer {token}")
    
    try:
        with urllib.request.urlopen(req) as response:
            with open(filename, "wb") as f:
                f.write(response.read())
        print(f"Successfully downloaded to {filename}!")
        
        # Đọc dữ liệu bằng openpyxl
        wb = openpyxl.load_workbook(filename, data_only=True)
        print("Tabs available in downloaded file:", wb.sheetnames)
        
        target_tab = "Hiệu quả bài (T3)"
        if target_tab in wb.sheetnames:
            sheet = wb[target_tab]
            print(f"\nReading first 100 rows from tab '{target_tab}':")
            rows = list(sheet.iter_rows(max_row=100, values_only=True))
            for i, row in enumerate(rows):
                clean_row = [str(cell) if cell is not None else "" for cell in row]
                if any(clean_row):
                    print(f"Row {i+1}: {clean_row[:15]}")
        else:
            print(f"Tab '{target_tab}' not found in workbook.")
            
        # Clean up
        if os.path.exists(filename):
            os.remove(filename)
            
    except urllib.error.HTTPError as e:
        print(f"HTTP Error: {e.code} - {e.reason}")
        # Thử đọc nội dung lỗi nếu có
        try:
            err_content = e.read().decode('utf-8')
            print("Error details:", err_content[:200])
        except Exception:
            pass
    except Exception as e:
        print(f"General Error: {e}")

if __name__ == "__main__":
    main()
